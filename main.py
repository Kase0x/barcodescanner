import openpyxl
from openpyxl import Workbook
from datetime import datetime
import os
import signal
import sys

# Define the filename
filename = 'inventory_log.xlsx'

# Dictionary to track inventory counts
inventory_counts = {}

class TimeoutError(Exception):
    pass

def timeout_handler(signum, frame):
    raise TimeoutError("Input timeout")

def load_inventory_from_excel():
    """Load existing inventory counts from Excel file"""
    global inventory_counts
    if os.path.exists(filename):
        wb = openpyxl.load_workbook(filename)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] and row[1] is not None:  # barcode and total count columns
                barcode = row[0]
                count = row[1]
                inventory_counts[barcode] = count

def get_operation():
    """Get the operation type from user"""
    while True:
        operation = input("Enter operation (ADD or REMOVE): ").strip().upper()
        if operation in ['ADD', 'REMOVE']:
            return operation
        print("Invalid operation. Please enter 'ADD' or 'REMOVE'.")

def get_barcode_with_timeout():
    """Get barcode input with 15-second timeout"""
    if os.name == 'nt':  # Windows
        import msvcrt
        import time
        
        print("Scan or enter barcode (15s timeout): ", end='', flush=True)
        start_time = time.time()
        barcode = ""
        
        while True:
            if msvcrt.kbhit():
                char = msvcrt.getch().decode('utf-8')
                if char == '\r':  # Enter key
                    print()
                    break
                elif char == '\b':  # Backspace
                    if barcode:
                        barcode = barcode[:-1]
                        print('\b \b', end='', flush=True)
                else:
                    barcode += char
                    print(char, end='', flush=True)
            
            if time.time() - start_time > 15:
                print("\nTimeout: No input received for 15 seconds.")
                return None
                
        return barcode.strip()
    else:  # Unix-like systems
        signal.signal(signal.SIGALRM, timeout_handler)
        signal.alarm(15)
        try:
            barcode = input("Scan or enter barcode (15s timeout): ").strip()
            signal.alarm(0)
            return barcode
        except TimeoutError:
            print("\nTimeout: No input received for 15 seconds.")
            return None

# Check if file exists, otherwise create it with headers
if not os.path.exists(filename):
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory"
    ws.append(["Barcode", "Total Count", "Last Updated"])
    wb.save(filename)

# Load existing inventory
load_inventory_from_excel()

# Open the workbook for appending
wb = openpyxl.load_workbook(filename)
ws = wb.active

print("Inventory management system started. Press Ctrl+C to stop.")

try:
    while True:  # Main operation loop
        # Get operation type
        operation = get_operation()
        print(f"Operation mode: {operation}")
        print("The system will return to operation selection after 15 seconds of no input.")
        
        # Barcode scanning loop for current operation
        while True:
            barcode = get_barcode_with_timeout()
            
            if barcode is None:  # Timeout occurred
                print("Returning to operation selection...\n")
                break  # Break inner loop to get new operation
            
            if barcode:
                timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                
                # Update inventory count
                current_count = inventory_counts.get(barcode, 0)
                if operation == 'ADD':
                    new_count = current_count + 1
                else:  # REMOVE
                    new_count = max(0, current_count - 1)
                
                inventory_counts[barcode] = new_count
                
                # Update or add barcode in Excel (find existing row or add new)
                barcode_found = False
                for row_num in range(2, ws.max_row + 1):
                    if ws.cell(row=row_num, column=1).value == barcode:
                        # Update existing row
                        ws.cell(row=row_num, column=2, value=new_count)
                        ws.cell(row=row_num, column=3, value=timestamp)
                        barcode_found = True
                        break
                
                if not barcode_found:
                    # Add new row for new barcode
                    ws.append([barcode, new_count, timestamp])
                
                wb.save(filename)
                
                print(f"Saved: {barcode} | {operation} | Count: {new_count} | Time: {timestamp}")
        
except KeyboardInterrupt:
    print("\nInventory system stopped.")
    wb.save(filename)
