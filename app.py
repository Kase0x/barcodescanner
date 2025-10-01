from flask import Flask, render_template, request, jsonify, send_file
from flask_sqlalchemy import SQLAlchemy
from flask_socketio import SocketIO, emit
from datetime import datetime, timedelta
import openpyxl
from openpyxl import Workbook
import os
import tempfile
import csv

app = Flask(__name__)
app.config['SECRET_KEY'] = 'your-secret-key-here'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///inventory.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)
socketio = SocketIO(app, cors_allowed_origins="*")

# Global variable to track last operation time for timeout
last_operation_time = None

# Global dictionary to store item descriptions from CSV
descriptions_cache = {}

class Inventory(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    barcode = db.Column(db.String(100), unique=True, nullable=False)
    total_count = db.Column(db.Integer, default=0, nullable=False)
    last_updated = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    description = db.Column(db.String(500), nullable=True)
    
    def to_dict(self):
        return {
            'id': self.id,
            'barcode': self.barcode,
            'total_count': self.total_count,
            'last_updated': self.last_updated.strftime("%Y-%m-%d %H:%M:%S"),
            'description': self.description or 'No description available'
        }

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/api/inventory')
def get_inventory():
    items = Inventory.query.all()
    return jsonify([item.to_dict() for item in items])

@app.route('/api/status')
def get_status():
    global last_operation_time
    
    if last_operation_time is None:
        time_remaining = 0
        requires_operation = True
    else:
        elapsed = datetime.utcnow() - last_operation_time
        time_remaining = max(0, 300 - elapsed.total_seconds())
        requires_operation = elapsed > timedelta(seconds=300)
    
    return jsonify({
        'last_operation_time': last_operation_time.isoformat() if last_operation_time else None,
        'time_remaining': time_remaining,
        'requires_operation': requires_operation
    })

@app.route('/api/scan', methods=['POST'])
def scan_barcode():
    global last_operation_time
    
    data = request.json
    barcode_input = data.get('barcode', '').strip().upper()
    operation_raw = data.get('operation')
    current_operation = operation_raw.upper() if operation_raw else 'ADD'
    
    if not barcode_input:
        return jsonify({'error': 'Barcode is required'}), 400
    
    # Check if input is an operation command
    if barcode_input in ['ADD', 'REMOVE']:
        last_operation_time = datetime.utcnow()
        socketio.emit('operation_changed', {
            'operation': barcode_input,
            'timestamp': last_operation_time.isoformat()
        })
        return jsonify({
            'success': True,
            'operation_changed': True,
            'new_operation': barcode_input,
            'message': f'Operation changed to {barcode_input}'
        })
    
    # Check if more than 300 seconds have passed since last operation
    if last_operation_time is None or (datetime.utcnow() - last_operation_time) > timedelta(seconds=300):
        return jsonify({
            'error': 'Operation timeout. Please enter ADD or REMOVE first.',
            'timeout': True,
            'requires_operation': True
        }), 400
    
    # Process as regular barcode
    barcode = barcode_input
    operation = current_operation
    
    if operation not in ['ADD', 'REMOVE']:
        return jsonify({'error': 'Operation must be ADD or REMOVE'}), 400
    
    # Find or create inventory item
    item = Inventory.query.filter_by(barcode=barcode).first()
    if not item:
        # Look up description from cache
        description = descriptions_cache.get(barcode, None)
        item = Inventory(barcode=barcode, total_count=0, description=description)
        db.session.add(item)
    elif not item.description and barcode in descriptions_cache:
        # Update existing item with description if it doesn't have one
        item.description = descriptions_cache[barcode]
    
    # Update count based on operation
    if operation == 'ADD':
        item.total_count += 1
    elif operation == 'REMOVE':
        item.total_count = max(0, item.total_count - 1)
    
    item.last_updated = datetime.utcnow()
    
    try:
        db.session.commit()
        
        # Update last operation time
        last_operation_time = datetime.utcnow()
        
        # Emit real-time update to all connected clients
        socketio.emit('inventory_update', item.to_dict())
        
        return jsonify({
            'success': True,
            'item': item.to_dict(),
            'operation': operation
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@app.route('/api/export')
def export_excel():
    try:
        # Create a new workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Inventory"
        
        # Add headers
        ws.append(["Barcode", "Total Count", "Last Updated", "Description"])
        
        # Add data
        items = Inventory.query.all()
        for item in items:
            ws.append([
                item.barcode,
                item.total_count,
                item.last_updated.strftime("%Y-%m-%d %H:%M:%S"),
                item.description or 'No description available'
            ])
        
        # Save to temporary file
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        wb.save(temp_file.name)
        temp_file.close()
        
        return send_file(
            temp_file.name,
            as_attachment=True,
            download_name=f'inventory_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/edit-quantity', methods=['POST'])
def edit_quantity():
    try:
        data = request.json
        barcode = data.get('barcode', '').strip()
        new_quantity = data.get('quantity')
        
        if not barcode:
            return jsonify({'error': 'Barcode is required'}), 400
        
        if new_quantity is None or not isinstance(new_quantity, int) or new_quantity < 0:
            return jsonify({'error': 'Valid quantity (non-negative integer) is required'}), 400
        
        # Find the inventory item
        item = Inventory.query.filter_by(barcode=barcode).first()
        if not item:
            return jsonify({'error': 'Item not found'}), 404
        
        # Update the quantity
        item.total_count = new_quantity
        item.last_updated = datetime.utcnow()
        
        db.session.commit()
        
        # Emit real-time update to all connected clients
        socketio.emit('inventory_update', item.to_dict())
        
        return jsonify({
            'success': True,
            'item': item.to_dict(),
            'message': f'Quantity updated for {barcode}'
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@app.route('/api/clear-database', methods=['POST'])
def clear_database():
    try:
        # Get confirmation from request
        data = request.json
        if not data or not data.get('confirmed'):
            return jsonify({'error': 'Confirmation required'}), 400
        
        # Clear all inventory entries
        deleted_count = Inventory.query.count()
        Inventory.query.delete()
        db.session.commit()
        
        # Emit update to all connected clients
        socketio.emit('database_cleared', {'deleted_count': deleted_count})
        
        return jsonify({
            'success': True,
            'message': f'Database cleared successfully. {deleted_count} entries removed.',
            'deleted_count': deleted_count
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

def load_descriptions():
    """Load item descriptions from InvDesc.csv into cache"""
    global descriptions_cache
    csv_file = 'instance/InvDesc.csv'
    
    if os.path.exists(csv_file):
        try:
            with open(csv_file, 'r', encoding='utf-8-sig', newline='') as file:
                csv_reader = csv.DictReader(file)
                for row in csv_reader:
                    item_id = row.get('Item ID', '').strip()
                    description = row.get('Item Description', '').strip()
                    if item_id and description:
                        descriptions_cache[item_id] = description
            print(f"Loaded {len(descriptions_cache)} descriptions from {csv_file}")
        except Exception as e:
            print(f"Error loading descriptions: {e}")
    else:
        print(f"Description file {csv_file} not found")

def migrate_database_schema():
    """Add description column to existing database if it doesn't exist"""
    try:
        # Check if description column exists
        with db.engine.connect() as conn:
            conn.execute(db.text("SELECT description FROM inventory LIMIT 1"))
        print("Description column already exists")
    except Exception:
        # Column doesn't exist, add it
        try:
            with db.engine.connect() as conn:
                conn.execute(db.text("ALTER TABLE inventory ADD COLUMN description VARCHAR(500)"))
                conn.commit()
            print("Added description column to database")
        except Exception as e:
            print(f"Error adding description column: {e}")

def migrate_existing_data():
    """Migrate data from existing Excel file to SQLite database"""
    excel_file = 'inventory_log.xlsx'
    if os.path.exists(excel_file):
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] and row[1] is not None:
                barcode = row[0]
                count = row[1]
                timestamp_str = row[2] if row[2] else datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                
                # Parse timestamp
                try:
                    if isinstance(timestamp_str, str):
                        timestamp = datetime.strptime(timestamp_str, "%Y-%m-%d %H:%M:%S")
                    else:
                        timestamp = timestamp_str
                except:
                    timestamp = datetime.now()
                
                # Check if item already exists
                existing_item = Inventory.query.filter_by(barcode=barcode).first()
                if not existing_item:
                    # Look up description from cache
                    description = descriptions_cache.get(barcode, None)
                    item = Inventory(
                        barcode=barcode,
                        total_count=count,
                        last_updated=timestamp,
                        description=description
                    )
                    db.session.add(item)
                elif not existing_item.description and barcode in descriptions_cache:
                    # Update existing item with description if it doesn't have one
                    existing_item.description = descriptions_cache[barcode]
        
        try:
            db.session.commit()
            print(f"Successfully migrated data from {excel_file}")
        except Exception as e:
            db.session.rollback()
            print(f"Error migrating data: {e}")

if __name__ == '__main__':
    with app.app_context():
        # Load descriptions first
        load_descriptions()
        db.create_all()
        # Migrate database schema to add description column if needed
        migrate_database_schema()
        migrate_existing_data()
    
    socketio.run(app, allow_unsafe_werkzeug=True , host='0.0.0.0', port=5000)